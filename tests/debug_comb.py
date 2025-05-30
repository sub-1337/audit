from PyQt6.QtWidgets import QApplication, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget
import openpyxl
import re
from enum import Enum

class MainWindow(QWidget):
   
    def getMergedCellVal(self, sheet, cell):
        rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
        return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value
    
    def unMergeCells(self, sheet_obj, processed):
        for i_row in range(1,self.rowMax):
            for i_col in range(1, self.colMax):              
                cell = sheet_obj.cell(row=i_row, column=i_col)
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    value_merged = self.getMergedCellVal(sheet_obj, cell)
                    processed[i_row][i_col] = value_merged             
                else:
                    processed[i_row][i_col] = sheet_obj.cell(row =  i_row, column = i_col).value
    
    def detectClass(self, string):
        if "пара" in string:
            return True
        else:
            return False
 
    def detectTime(self, string):
        res = re.search("[+-]?([0-9]*[.])?[0-9]+?\/[+-]?([0-9]*[.])?[0-9]+", string)
        if res:
            return True
        else:
            return False
    
    def detectPartOfGroup(self, string):
        if "п/гр" in string:
            return True
        else:
            return False
    class DirectionOfGroupingUp(Enum):
        TOP = 0
        BOTTOM = 1
        RIGHT = 2
        LEFT = 3
    
    def groupUpBorders(self, sheet_obj, processed, direction : DirectionOfGroupingUp):
        def check(i_row_parent, i_col_parent, i_row, i_col, is_going, sheet_obj, processed, direction : self.DirectionOfGroupingUp):
            def checkAny(cell):
                if cell.border.top.style != None or \
                    cell.border.bottom.style != None or \
                    cell.border.left.style != None or \
                    cell.border.right.style != None:
                    return True
                else:
                    return False
            def checkAll(cell):
                if cell.border.top.style != None and \
                    cell.border.bottom.style != None and \
                    cell.border.left.style != None and \
                    cell.border.right.style != None:
                    return True
                else:
                    return False
            def checkSide(cell, direction : self.DirectionOfGroupingUp):
                def checkTop(cell):
                    if cell.border.top.style == None:
                        return True
                    else:
                        return False
                def checkBottom(cell):
                    if cell.border.bottom.style == None:
                        return True
                    else:
                        return False
                if direction == self.DirectionOfGroupingUp.TOP:
                    return checkTop(cell)
                elif direction == self.DirectionOfGroupingUp.BOTTOM:
                    return checkBottom(cell)
            
            cell = sheet_obj.cell(row=i_row, column=i_col)
            origin_cell = sheet_obj.cell(row=i_row_parent, column=i_col_parent)
            if checkAny(cell):
                if checkAll(cell) == False:
                    if cell.value != None:
                        if not self.detectClass(cell.value):
                            if checkSide(cell, direction):
                                if self.detectTime(cell.value) == False:
                                    if self.detectPartOfGroup(cell.value) == False:
                                        if i_row < self.rowMax and  i_col < self.colMax:                                                
                                            if direction == self.DirectionOfGroupingUp.TOP:
                                                if (i_row - 1) > 1:
                                                    processed[i_row - 1][i_col] = origin_cell.value
                                                    is_going = True
                                                    check(i_row_parent, i_col_parent, i_row - 1, i_col, is_going, sheet_obj, processed, direction)
                                            if direction == self.DirectionOfGroupingUp.BOTTOM:
                                                if (i_row + 1) < self.rowMax:                                        
                                                    processed[i_row + 1][i_col] = origin_cell.value
                                                    is_going = True
                                                    check(i_row_parent, i_col_parent, i_row + 1, i_col, is_going, sheet_obj, processed, direction)
                    else: #cell.value == None:
                        if checkSide(cell, direction):
                            if is_going:
                                if i_row < self.rowMax and  i_col < self.colMax:                                                                                    
                                    if direction == self.DirectionOfGroupingUp.TOP:
                                        if (i_row - 1) > 1:
                                            processed[i_row - 1][i_col] = origin_cell.value
                                            check(i_row_parent, i_col_parent, i_row - 1, i_col, is_going, sheet_obj, processed, direction)
                                    if direction == self.DirectionOfGroupingUp.BOTTOM:
                                        if (i_row + 1) < self.rowMax:                               
                                            processed[i_row + 1][i_col] = origin_cell.value
                                            check(i_row_parent, i_col_parent, i_row + 1, i_col, is_going, sheet_obj, processed, direction)                  
        for i_row in range(1,self.rowMax):
            for i_col in range(1, self.colMax):
                check(i_row, i_col, i_row, i_col, False, sheet_obj, processed, direction)

        
    def __init__(self):
        super().__init__()

        path = "D:\\git\\audit\\data\\1kurs.xlsx"
        self.rowMax = 100
        self.colMax = 25
        # To open the workbook 
        # workbook object is created
        wb_obj = openpyxl.load_workbook(path)

        # Get workbook active sheet object
        # from the active attribute
        sheet_obj = wb_obj.active

        self.setWindowTitle("Пример QTableWidget")
        self.resize(600, 400)

        # Создание таблицы
        self.table = QTableWidget()
        self.table.setRowCount(self.rowMax)  # Количество строк
        self.table.setColumnCount(self.colMax)  # Количество столбцов
        #self.table.setHorizontalHeaderLabels(["Имя", "Возраст", "Город"])  # Заголовки столбцов
        
        # Избавится от объеденённых ячеек
        processed = [ [0]*self.colMax for i in range(self.rowMax)]
        self.unMergeCells(sheet_obj, processed)
        self.groupUpBorders(sheet_obj, processed, self.DirectionOfGroupingUp.TOP)
        self.groupUpBorders(sheet_obj, processed, self.DirectionOfGroupingUp.BOTTOM)


        for i_row in range(1,self.rowMax):
            for i_col in range(1, self.colMax):
                item = QTableWidgetItem(processed[i_row][i_col])
                self.table.setItem(i_row - 1, i_col - 1, item)

        # Компоновка
        layout = QVBoxLayout()
        layout.addWidget(self.table)
        self.setLayout(layout)

app = QApplication([])
window = MainWindow()
window.show()
app.exec()