import re
from enum import Enum
import openpyxl
from openpyxl import Workbook
from core.data_model import InputData
import core.data_model as dm
import datetime as dt

class DocumentReader():
    class DirectionOfGroupingUp(Enum):
        """
        В какую сторону раскрывать поля (не используется)
        """
        TOP = 0
        BOTTOM = 1
        RIGHT = 2
        LEFT = 3

    def getMergedCellVal(self, sheet, cell):
        """
        Получить значение объединённой ячейки
        """
        rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
        return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value
    
    def unMergeCells(self, sheet_obj, processed):
        """
        Раскрыть объединённые ячейки
        """
        for i_row in range(1,self.data.rowMax):
            for i_col in range(1, self.data.colMax):              
                cell = sheet_obj.cell(row=i_row, column=i_col)
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    value_merged = self.getMergedCellVal(sheet_obj, cell)
                    processed[i_row][i_col] = value_merged             
                else:
                    processed[i_row][i_col] = sheet_obj.cell(row =  i_row, column = i_col).value
    
    def detectClass(self, string):
        """
        Детект слова "пара" в ячейке
        """
        if isinstance(string, str):
            if "пара" in string:
                return True
            else:
                return False
        else:
            return False
 
    def detectTime(self, string):
        """
        Детект значения времени в ячейке
        """
        res = re.search("[+-]?([0-9]*[.])?[0-9]+?\/[+-]?([0-9]*[.])?[0-9]+", string)
        if res:
            return True
        else:
            return False
    
    def groupUpAndMerge(self, sheet_obj, processed):
        """
        Дублировать объединённые ячейки, обрамлённые одним бордюром
        """
        def check(i_row_parent, i_col_parent, i_row, i_col, is_going, sheet_obj, processed, direction : self.DirectionOfGroupingUp, groupIndexes = []):
            """
            Проверить ячейку на предмет единого бордюра
            """
            def checkAny(cell):
                """
                Есть ли хотя бы одна сторона не ограниченная бордюром
                """
                if cell.border.top.style != None or \
                    cell.border.bottom.style != None or \
                    cell.border.left.style != None or \
                    cell.border.right.style != None:
                    return True
                else:
                    return False
            def checkAll(cell):
                """
                Проверить что клетка полностью обрамлена бордюрами
                """
                if cell.border.top.style != None and \
                    cell.border.bottom.style != None and \
                    cell.border.left.style != None and \
                    cell.border.right.style != None:
                    return True
                else:
                    return False
            def checkSide(cell, direction : self.DirectionOfGroupingUp):
                """
                Проверить сторону на предмет бордюра
                """
                def checkTop(cell):
                    if cell.border.top.style == None:
                        return True
                    else:
                        return False
                def checkBottom(cell):
                    if cell.border.bottom.style == None:
                        return True
                    else:
                        return False
                if direction == self.DirectionOfGroupingUp.TOP:
                    return checkTop(cell)
                elif direction == self.DirectionOfGroupingUp.BOTTOM:
                    return checkBottom(cell)
            
            cell = sheet_obj.cell(row=i_row, column=i_col)
            if checkAny(cell):
                if checkAll(cell) == False:
                    if cell.value != None:
                        if not self.detectClass(cell.value):
                            if checkSide(cell, direction):
                                if self.detectTime(cell.value) == False:                                    
                                    if i_row < self.data.rowMax and  i_col < self.data.colMax:
                                        if direction == self.DirectionOfGroupingUp.BOTTOM:
                                            if (i_row + 1) < self.data.rowMax:
                                                # Добавить клеточку в список объединяемых
                                                groupIndexes.append((i_row,i_col,))
                                                #  Рекурсивный вызов для новой клеточки (ниже чем текущая)
                                                check(i_row_parent, i_col_parent, i_row + 1, i_col, is_going, sheet_obj, processed, direction, groupIndexes)
                            else:
                                # Добавить текущую клеточку (которая упирается в бордюр)
                                groupIndexes.append((i_row,i_col,))
                    else: #cell.value == None:
                        if checkSide(cell, direction):
                            if i_row < self.data.rowMax and  i_col < self.data.colMax:                                                                                    
                                if direction == self.DirectionOfGroupingUp.BOTTOM:
                                    if (i_row + 1) < self.data.rowMax: 
                                        # Добавить текущую пустую клеточку в список объединяемых
                                        groupIndexes.append((i_row,i_col,))
                                        # Рекурсивно проверить следующую клеточку (ниже чем текущая)
                                        check(i_row_parent, i_col_parent, i_row + 1, i_col, is_going, sheet_obj, processed, direction, groupIndexes)
                        else:
                            # Добавить текущую клеточку (которая упирается в бордюр)
                            groupIndexes.append((i_row,i_col,))
            # Вернуть список объединяемых клеточек
            return groupIndexes  
        
        def groupUp(group):
            """
            Сгруппировать уникальные значения ячеек
            """

            uniqueValues = []
            for cellIndex in group:
                currCell = processed[cellIndex[0]][cellIndex[1]]
                if currCell:
                    if currCell not in uniqueValues:
                        uniqueValues.append(currCell)

            cellSummed = ''

            for cellCurr in uniqueValues:
                cellSummed += ' '
                cellSummed += cellCurr
            
            for cellIndex in group:
                processed[cellIndex[0]][cellIndex[1]] = cellSummed
        visited = []
        for i_row in range(1,self.data.rowMax):
            for i_col in range(1, self.data.colMax):
                if not ((i_row, i_col,) in visited):
                    currentGroup = check(i_row, i_col, i_row, i_col, False, sheet_obj, processed, self.DirectionOfGroupingUp.BOTTOM, [])
                    if len(currentGroup) > 0:
                        groupUp(currentGroup)
                    visited += currentGroup
                    pass
                else:
                    next
            

    def filterSpaces(self):
        """
        Убрать лишние пробелы
        """
        for i_row in range(1,self.data.rowMax):
            for i_col in range(1, self.data.colMax):
                blockStr = self.data.processed[i_row][i_col]
                if (not blockStr) or isinstance(blockStr, int):
                    blockStr = ""
                self.data.processed[i_row][i_col] = " ".join(blockStr.split())

    def readDoc(self, dayOfStartDic):
        """
        Прочитать документ и составить все таблицы
        """
        if self.worbookNamesCurrent:
            sheet_obj = self.wb_obj[self.worbookNamesCurrent]
        else:
            sheet_obj = self.wb_obj.active
        
        self.dayOfStartDic = dayOfStartDic

        # Создать рабочую структуру
        self.data.processed = [ [""]*self.data.colMax for i in range(self.data.rowMax)]
        # Избавится от объеденённых ячеек

        self.unMergeCells(sheet_obj, self.data.processed)

        # Склеить ячейки в одном блоке бордюра
        self.groupUpAndMerge(sheet_obj, self.data.processed)

        # Убрать лишние пробелы
        self.filterSpaces()

        # Распарсить данные таблицы сверху/слева такие как группа и время
        self.parseBorders()

        # Распарсить правила
        self.parseData()

        # Создать из данных таблицу по дням
        self.calcYear()
        # Создать таблицу из данных по аудиториям
        self.calcAuditories()
    def readHead(self):
        """
        Прочитать заголовок эксель файла на предмет книг
        """       
        self.wb_obj = openpyxl.load_workbook(self.docPath)
        self.worbookNames = self.wb_obj.sheetnames
    def __init__(self, docPath):
        super().__init__()
        # Данные распаршенные напрямую
        self.data = InputData()
        # Данные по дням
        self.dataYear = dm.CalenderYear()
        # Список всех правил
        self.rules = dm.Rules()
        # День начала учебного года
        self.dayOfStartDic = None
        # Длительность учебного года
        self.maxWeek = 17
        # Все аудитории
        self.allAuditories = dm.AllAuditories()

        self.data.rowMax = 100
        self.data.colMax = 100
        # Путь до файла
        self.docPath = docPath

        # Текущий вариант книги эксель
        self.worbookNamesCurrent = None
        self.readHead()        
        
    def isWeekDayName(self, cell):
        """
        Является ли строка днём недели
        """
        if cell == 'п о н е д е л ь н и к' or cell == 'П о н е д е л ь н и к' or cell == 'понедельник':
            return dm.DayOfWeek.MONDAY
        elif cell == 'в т о р н и к' or cell == 'В т о р н и к' or cell == 'вторник':
            return dm.DayOfWeek.TUESDAY
        elif cell == 'с р е д а' or cell == 'С р е д а' or cell == 'среда':
            return dm.DayOfWeek.WEDNESDAY
        elif cell == 'ч е т в е р г' or cell == 'Ч е т в е р г' or cell == 'четверг':
            return dm.DayOfWeek.THURSDAY
        elif cell == 'п я т н и ц а' or cell == 'П я т н и ц а' or cell == 'пятница':
            return dm.DayOfWeek.FRIDAY
        elif cell == 'с у б б о т а' or cell == 'С у б б о т а' or cell == 'суббота':
            return dm.DayOfWeek.SATURDAY
        else:
            return None
    def isPara(self, cell):
        """
        Является ли строка - парой
        """
        if 'пара' in cell:
            number = cell[0]
            number = int(number)
            return dm.Para(number)
        else:
            return None
    def isTime(self, cell):
        """
        Является ли строка временем занятия
        """
        pattern = r'\d{1,2}\.\d{2}/\d{1,2}\.\d{2}'
        match = re.search(pattern, cell)
        if match:
            splitted = cell.split('/')
            for curr in splitted:
                s = curr.split('.')
                return dt.time(int(s[0]),int(s[1]))
        else:
            return None
    @staticmethod
    def isGroup(cell):
        """
        Является ли строка названием группы
        """
        groupNames = ['ИВТ', 'ПМ', 'ИСТ', 'КТЭС', 'Р', 'ИТС', 'РЭС',\
                      'СИБ', 'ССК', 'ТР', 'ИТД', 'КТ', 'СБК', 'СТ', 'ИС',\
                      'ПО', 'ВМ', 'ИС']
        for name in groupNames:
            if name in cell:
                return True
        return False 
    @staticmethod
    def isIgnoreTopWords(cell):
        """
        Является ли игнормруемой строкой сверху
        """
        if ('Р А С П И С А Н И Е' in cell) or ('РАСПИСАНИЕ' in cell) or  \
           ('Расписание' in cell) or ('расписание' in cell):
            return True
        else:
            return False
    def parseBorders(self): 
        """
        Распарсить данные таблицы сверху/слева такие как группа и время
        """

        # Пройтись по левой колонке где день и пара/время
        i_row = 0
        readed = False
        self.leftColumnData = []
        while not readed:
            for j_col in range(15):
                cell = self.data.processed[i_row][j_col]
                weekday = self.isWeekDayName(cell)
                if weekday:
                    cellRightNear = self.data.processed[i_row][j_col + 1]
                    para = self.isPara(cellRightNear)
                    if para:
                        cellRightBottomNear = self.data.processed[i_row + 1][j_col + 1]
                        if self.isTime(cellRightBottomNear):
                            # Сдесь мы в левой верхней границе
                            #if (self.isPara(cell) == False) and (self.isWeekDayName(cell) == None):
                            self.leftColumnData.append({'weekday' : weekday, 'para' : para, 'rows' : (i_row, i_row + 1,)})
            if i_row + 1 < self.data.rowMax:
                i_row += 1
            else:
                readed = True

        # Пройтись по верху где группа
        j_col = 0
        readed = False
        self.topRowData = []
        while not readed:
            for i_row in range(30):
                cell = self.data.processed[i_row][j_col]
                if cell:
                    if self.isGroup(cell):
                        if not self.isIgnoreTopWords(cell):
                            self.topRowData.append({'group' : dm.Group(cell), 'col' : j_col})

            if j_col + 1 < self.data.colMax:
                j_col += 1
            else:
                readed = True

        #self.rules.addRule(dm.Rule(dm.Auditory("5442"), dm.DayOfWeek.MONDAY, dm.Para(1), dm.RuleEven.DEFAULT, None, dm.RuleSubgroup.DEFAULT, dm.Id(1)))
    @staticmethod
    def parseCell(cell):
        """
        Прочитать ячейку и вернуть заготовку для правила
        """
        if not cell:
            return None
        if "MOODLE" in cell:
            return None
        
        cloneResult  = {'auditory' : None, 'even' : None, 'week' : None, 'subgroup' : None, 'confidence' : None}
        result = []

        confidence = 100

        auditory = dm.Auditory(None)
        auditory2 = dm.Auditory(None)
        even = dm.RuleEven.DEFAULT
        week = []
        subgroup = dm.RuleSubgroup.DEFAULT
        subgroup2 = None

        if ('чн.' in cell) or ('чн' in cell):
            even = dm.RuleEven.EVEN
            if ('нч.' in cell) or ('нч' in cell):
                confidence -= 20
            
        if ('нч.' in cell) or ('нч' in cell):
            even = dm.RuleEven.ODD
            if ('чн.' in cell) or ('чн' in cell):
                confidence -= 20
        
        comment = cell.replace('чн.','')
        comment = comment.replace('нч.','')

        patternAud = r'[0-9]{4}'
        matches = re.findall(patternAud, cell)
        comment = re.sub(patternAud, "", comment)
        if len(matches) != 1:
            if len(matches) == 2:
                auditory = dm.Auditory(int(matches[0]))
                auditory2 = dm.Auditory(int(matches[1]))
            else:
                confidence -= 20
        else:
            auditory = dm.Auditory(int(matches[0]))

        countOfSubgroup = 0
        countOfSubgroup += cell.count('подгр')
        countOfSubgroup += cell.count('п/гр')
        
        comment = comment.replace("а.", "")
        
        if countOfSubgroup > 1:
            subgroupRegexp = r"(\d)\s*(?:п/гр\.|подгр)\s*-?\s*([\d,]+)"
            matchesSubGroup = re.findall(subgroupRegexp, cell)
            comment = re.sub(subgroupRegexp, "", comment)
            
            comment = comment.replace('нед', '')
            comment = re.sub(r"\s+", " ", comment).strip()
            
            if len(matchesSubGroup) > 1:
                result.append({'auditory' : auditory, 'even' : even, 'week' : week, 'subgroup' : subgroup, 'confidence' : confidence, 'comment' : comment})
                for match in matchesSubGroup:
                    result.append(result[0].copy())
                    result[-1]['subgroup'] = dm.RuleSubgroup(int(match[0]))
                    result[-1]['week'] = [int(x) if x != '' else 0 for x in match[1].split(',')]
                del result[0]
            else:
                result.append({'auditory' : auditory, 'even' : even, 'week' : week, 'subgroup' : subgroup, 'confidence' : confidence, 'comment' : comment})
        else:
            #subgroupRegexp = r'(\d+)\s*(?=п/гр\.|подгр)'
            subgroupRegexp = r'\s(\d)\b'
            subgroupMatch = re.findall(subgroupRegexp, cell)
            if len(subgroupMatch) == 2:
                subgroupNumber = int(subgroupMatch[0])
                subgroupNumber2 = int(subgroupMatch[1])
                if (subgroupNumber < 0 or subgroupNumber > 4) or \
                   (subgroupNumber2 < 0 or subgroupNumber2 > 4):
                    confidence -= 20
                    subgroup = dm.RuleSubgroup.DEFAULT
                    subgroup2 = dm.RuleSubgroup.DEFAULT
                else:
                    subgroup = dm.RuleSubgroup(subgroupNumber)
                    subgroup2 = dm.RuleSubgroup(subgroupNumber2)
            elif len(subgroupMatch) == 1:
                subgroupNumber = int(subgroupMatch[0])
                if subgroupNumber < 0 or subgroupNumber > 4:
                    confidence -= 20
                    subgroup = dm.RuleSubgroup.DEFAULT
                else:
                    subgroup = dm.RuleSubgroup(subgroupNumber)
                
            else:
                subgroup = dm.RuleSubgroup.DEFAULT
            if subgroup2 and auditory2:
                result.append({'auditory' : auditory, 'even' : even, 'week' : week, 'subgroup' : subgroup, 'confidence' : confidence, 'comment' : comment})
                result.append({'auditory' : auditory2, 'even' : even, 'week' : week, 'subgroup' : subgroup2, 'confidence' : confidence, 'comment' : comment})
            else:
                result.append({'auditory' : auditory, 'even' : even, 'week' : week, 'subgroup' : subgroup, 'confidence' : confidence, 'comment' : comment})
        return result

    def parseData(self):
        """
        Распарсить правила
        """
        self.rules = dm.Rules()
        idNum = 1
        for left in self.leftColumnData:
            for top in self.topRowData:
                weekday = left['weekday']
                para = left['para']
                rows = left['rows']
                
                group = top['group']
                col = top['col']

                res1 = self.parseCell(self.data.processed[rows[0]][col])
                res2 = self.parseCell(self.data.processed[rows[1]][col])

                resTotal = []
                if res1:
                    resTotal += res1
                if res2:
                    resTotal += res2

                for rule in resTotal:
                    if rule:
                        self.rules.addRule(dm.Rule(rule['auditory'], weekday, para,  rule['even'],\
                                                   rule['week'], rule['subgroup'], rule['comment'], rule['confidence'], dm.Group(group)))
        pass

    def calcAuditories(self):
        """
        Создать таблицу по аудиториям
        """

        auditories = {}

        for day in self.dataYear.allDays:
            for block in day.blocks:
                #print(block)
                if not auditories.get(block.auditory):
                    auditories[block.auditory] = []
                auditories[block.auditory].append(block)
            pass

        self.allAuditories.auditories = auditories
        self.allAuditories.sort()
        
    def calcYear(self):
        """
        Создать из данных таблицу по дням
        """
        startDate = dt.datetime(self.dayOfStartDic['year'], self.dayOfStartDic['month'], self.dayOfStartDic['day'])
        endDate = startDate + dt.timedelta(days = 31 * self.maxWeek)
        currentDate = startDate
        weekStart = startDate.isocalendar().week
        weekNumber = 0
        weekDay = 0

        while currentDate <= endDate:
            weekNumber = currentDate.isocalendar().week - weekStart            
            weekDay = currentDate.weekday()
            #(currentDate.strftime("%Y-%m-%d") + ' ' + f"week: {weekNumber} weekday {weekDay}")            
            
            if weekNumber > self.maxWeek:
                return

            dayOfWeekEnum = dm.DayOfWeek(weekDay)
            if weekDay % 2 == 0:
                even = dm.RuleEven.EVEN
            else:
                even = dm.RuleEven.ODD

            day = dm.CalenderDay(dt.date(currentDate.year, currentDate.month, currentDate.day))
            dayDate =  dt.date(currentDate.year, currentDate.month, currentDate.day)
            for rule in self.rules.rules:
                if rule.dayOfWeek == dayOfWeekEnum:
                    if len(rule.week) == 0:
                        if rule.even != dm.RuleEven.DEFAULT:
                            if rule.even == even:
                                block = dm.CalenderBlock(None, None, rule.para, rule.auditory, rule.subgroup, rule.comment, dayDate, rule.group)
                                day.addBlock(block)
                        else:
                            block = dm.CalenderBlock(None, None, rule.para, rule.auditory, rule.subgroup, rule.comment, dayDate, rule.group)
                            day.addBlock(block)
                    else:
                        if weekNumber in rule.week:
                            block = block = dm.CalenderBlock(None, None, rule.para, rule.auditory, rule.subgroup, rule.comment, dayDate, rule.group)
                            day.addBlock(block) 
            self.dataYear.addDay(day)
            currentDate += dt.timedelta(days=1)


        # day = dm.CalenderDay(dt.date(2025,11,10))
        # block = dm.CalenderBlock(None, dt.time(11,30), dm.Para(1), dm.Auditory(1000))
        #day.addBlock(block)
        #self.dataYear.addDay(day)
    def getRules(self):
        """
        Вернуть правила
        """
        return self.rules
    
    def writeReportAuditory(self, auditoryBlocks, fileName):
        """
        Распечатать по аудитории файл
        """
        wb = Workbook()
        ws = wb.active
        # ws.cell(10, 5).value = "TEST"
        firstRow = 1

        ws.cell(firstRow, 1).value = "Аудитория"
        ws.cell(firstRow, 2).value = "Дата"
        ws.cell(firstRow, 3).value = "Пара"
        ws.cell(firstRow, 4).value = "Группа"
        ws.cell(firstRow, 5).value = "Подгруппа"
        ws.cell(firstRow, 6).value = "Комментарий"

        # dataFirstRow = firstRow + 1

        rowsCount = len(auditoryBlocks) # Массив 
        colsCount = 5

        # Итерация по массиву с первым элемиентом 0
        # Далле вычисляем позицию в документе по смещению
        # Т.к. в жокументе пнрвый элемент 1
        for i_row_array in range(0, rowsCount):
            i_row_document = i_row_array + 2
            ws.cell(i_row_document, 1).value = str(auditoryBlocks[i_row_array].auditory)
            ws.cell(i_row_document, 2).value = str(auditoryBlocks[i_row_array].date)
            ws.cell(i_row_document, 3).value = str(auditoryBlocks[i_row_array].para)
            ws.cell(i_row_document, 4).value = str(auditoryBlocks[i_row_array].groupBase)            
            ws.cell(i_row_document, 5).value = str(auditoryBlocks[i_row_array].subGroup)
            ws.cell(i_row_document, 6).value = str(auditoryBlocks[i_row_array].comment)

        # Save the workbook
        wb.save(fileName)

    def writeReportDay(self, day : dm.CalenderDay, fileName):
        """
        Распечатать по дню файл
        """
        
        wb = Workbook()
        ws = wb.active
        firstRow = 1

        ws.cell(firstRow, 1).value = "Аудитория"
        for i in range(1,7):
            ws.cell(firstRow, 1 + i).value = f"Пара {i}"

        auditorArr = day.ReturnArrayByAuditory()
        rowMax = len(auditorArr)
        colMax = len(auditorArr[0])

        def GetCell(block : dm.CalenderBlock):
            return f"группа {block.groupBase} подгруппа {block.subGroup}\n{block.comment}"

        # Итерация по массиву с первым элемиентом 0
        # Далле вычисляем позицию в документе по смещению
        # Т.к. в жокументе пнрвый элемент 1
        for i_row_array in range(0, rowMax):
            i_row_document = i_row_array + 2
            for j_col in range(colMax):
                j_col_document = j_col + 1
                if isinstance(auditorArr[i_row_array][j_col], dm.CalenderBlock):
                    ws.cell(i_row_document, j_col_document).value = GetCell(\
                        auditorArr[i_row_array][j_col])
                else:
                    ws.cell(i_row_document, j_col_document).value = str(\
                        auditorArr[i_row_array][j_col])
            # Изменить размер колонок
            colNum = i_row_document
            colLetter = openpyxl.utils.get_column_letter(colNum)
            ws.column_dimensions[colLetter].width = 60
            

        wb.save(fileName)
        pass